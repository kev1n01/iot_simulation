from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import pandas as pd

# Load the uploaded CSV file to check the data structure
file_path = 'datos_simulados_sistema_acuaponico.csv'
data = pd.read_csv(file_path)

# Display the first few rows to understand its structure
data.head()

# Define allowable ranges and recommendations for each parameter
decisions_data = {
    "Parametro": [
        "Nivel de Oxígeno en Agua (mg/L)", "Nivel de pH", "Nivel de Nitratos (ppm)",
        "Nivel de Nitritos (ppm)", "Temperatura del Agua (°C)", "Temperatura Ambiente (°C)",
        "Humedad Ambiente (%)", "Cantidad de Alimento (g)", "Flujo de Agua (L/min)",
        "Intensidad de Luz (lux)", "Nivel de Agua (cm)", "Estado del Filtro", "Consumo de Energía (kWh)"
    ],
    "Valor Minimo": [
        5.0, 6.5, 10, 0.5, 18, 10, 40, 50, 5, 5000, 20, None, 1
    ],
    "Valor Maximo": [
        8.0, 7.5, 50, 1.0, 28, 35, 70, 100, 10, 80000, 80, None, 10
    ],
    "Accion a Tomar": [
        "Ajustar bomba de oxígeno", "Ajustar sistema de pH", "Revisar nivel de filtrado",
        "Ajustar control de nitritos", "Ajustar temperatura", "Ajustar ventilación",
        "Ajustar humedad", "Revisar nivel de alimentación", "Ajustar flujo de agua",
        "Ajustar intensidad de luz", "Ajustar nivel de agua", "Cambiar filtro si está sucio",
        "Revisar consumo de energía"
    ]
}

# Create a new workbook and select active sheet
wb = Workbook()
ws = wb.active
ws.title = "Parametros y Ajustes"

# Add a title
ws.merge_cells("A1:D1")
title_cell = ws["A1"]
title_cell.value = "Parámetros del Sistema Acuapónico y Acciones"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal="center")

# Convert decisions data to a DataFrame and add to the sheet
decisions_df = pd.DataFrame(decisions_data)

# Write headers with formatting
headers = ["Parametro", "Valor Minimo", "Valor Maximo", "Accion a Tomar"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_num, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Write data rows
for row in dataframe_to_rows(decisions_df, index=False, header=False):
    ws.append(row)

# Apply border to data cells
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Add autofilter to columns
ws.auto_filter.ref = "A2:D" + str(ws.max_row)

# Add logo if available
logo_path = 'logo.jpg'
try:
    logo = Image(logo_path)
    ws.add_image(logo, "E1")
except FileNotFoundError:
    pass  # Skip if no logo is provided

# Save the workbook
output_path = "./data/decision_de_parametros_Sistema_Acuaponico.xlsx"
wb.save(output_path)

output_path
